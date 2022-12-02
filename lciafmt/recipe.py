# recipe.py (lciafmt)
# !/usr/bin/env python3
# coding=utf-8
"""
This module contains functions needed to compile LCIA methods from the
ReCiPe model
"""

import pandas as pd
import openpyxl

import lciafmt.cache as cache
import lciafmt.df as dfutil
import lciafmt.xls as xls

from .util import datapath, aggregate_factors_for_primary_contexts, log,\
        format_cas


contexts = {
        'urban air': 'air/urban',
        'urban  air': 'air/urban',
        'Urban air': 'air/urban',
        'Rural air': 'air/rural',
        'rural air': 'air/rural',
        'agricultural soil': 'soil/agricultural',
        'Agricultural soil': 'soil/agricultural',
        'industrial soil': 'soil/industrial',
        'Industrial soil': 'soil/industrial',
        'freshwater': 'water/freshwater',
        'Freshwater': 'water/freshwater',
        'fresh water': 'water/freshwater',
        'seawater': 'water/sea water',
        'sea water': 'water/sea water',
        'Sea water': 'water/sea water',
        'marine water': 'water/sea water'}
flowables_split = pd.read_csv(datapath + 'ReCiPe2016_split.csv')


def get(add_factors_for_missing_contexts=True, endpoint=True,
        summary=False, file=None, url=None) -> pd.DataFrame:
    """Generate a method for ReCiPe 2016 in standard format.

    :param add_factors_for_missing_contexts: bool, if True generates average
        factors for unspecified contexts
    :param endpoint: bool, if True generates endpoint indicators from midpoints
    :param summary: bool, if True aggregates endpoint methods into
        summary indicators
    :param file: str, alternate filepath for method, defaults to file stored
        in cache
    :param url: str, alternate url for method, defaults to url in method config
    :return: DataFrame of method in standard format
    """
    log.info("getting method ReCiPe 2016")
    f = file
    if f is None:
        fname = "recipe_2016.xlsx"
        if url is None:
            url = ("http://www.rivm.nl/sites/default/files/2018-11/" +
                   "ReCiPe2016_CFs_v1.1_20180117.xlsx")
        f = cache.get_or_download(fname, url)
    df = _read(f)
    if add_factors_for_missing_contexts:
        log.info("adding average factors for primary contexts")
        df = aggregate_factors_for_primary_contexts(df)

    if endpoint:
        endpoint_df, endpoint_df_by_flow = _read_endpoints(f)
        log.info("converting midpoints to endpoints")
        # first assesses endpoint factors that are specific to flowables
        flowdf = df.merge(endpoint_df_by_flow, how="inner",
                          on=["Method", "Flowable"])
        flowdf.rename(columns={'Indicator_x': 'Indicator',
                               'Indicator_y': 'EndpointIndicator'},
                      inplace=True)
        # next apply endpoint factors by indicator
        df2 = df.merge(endpoint_df, how="inner", on=["Method", "Indicator"])
        df2 = pd.concat([df2, flowdf], ignore_index=True, sort=False)
        # reformat dataframe and apply conversion
        df2['Characterization Factor'] = df2['Characterization Factor'] * df2['EndpointConversion']
        df2['Method'] = df2['EndpointMethod']
        df2['Indicator'] = df2['EndpointIndicator']
        df2['Indicator unit'] = df2['EndpointUnit']
        df2.drop(columns=['EndpointMethod', 'EndpointIndicator',
                          'EndpointUnit', 'EndpointConversion'],
                 inplace=True)
        df = pd.concat([df, df2], ignore_index=True, sort=False)

    log.info("handling manual replacements")
    """due to substances listed more than once with the same name but
    different CAS, this replaces all instances of the Original Flowable with
    a New Flowable based on a csv input file according to the CAS"""
    for index, row in flowables_split.iterrows():
        newCAS = format_cas(row['CAS'])
        newFlow = row['New Flowable']
        df.loc[df['CAS No'] == newCAS, 'Flowable'] = newFlow

    length = len(df)
    df.drop_duplicates(keep='first', inplace=True)
    length = length - len(df)
    log.info(f"{length} duplicate entries removed")

    if summary:
        log.info("summarizing endpoint categories")
        endpoint_categories = df.groupby(['Method', 'Method UUID',
                                          'Indicator unit', 'Flowable',
                                          'Flow UUID', 'Context', 'Unit',
                                          'CAS No', 'Location',
                                          'Location UUID', 'EndpointCategory'],
                                         as_index=False)['Characterization Factor'].sum()
        endpoint_categories['Indicator'] = endpoint_categories['EndpointCategory']
        endpoint_categories['Indicator UUID'] = ""
        endpoint_categories.drop(columns=['EndpointCategory'], inplace=True)

        # To append endpoint categories to exisiting endpointLCIA,
        # set append = True, otherwise replaces endpoint LCIA
        append = False
        if append:
            log.info("appending endpoint categories")
            df = pd.concat([df, endpoint_categories], sort=False)
        else:
            log.info("applying endpoint categories")
            df = endpoint_categories

        # reorder columns in DF
        df = df.reindex(columns=["Method", "Method UUID", "Indicator",
                                 "Indicator UUID", "Indicator unit", "Flowable",
                                 "Flow UUID", "Context", "Unit", "CAS No",
                                 "Location", "Location UUID",
                                 "Characterization Factor"])
    return df


def _read(file: str) -> pd.DataFrame:
    log.info(f"read ReCiPe 2016 from file {file}")
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    records = []
    for name in wb.sheetnames:
        if _eqstr(name, "Version") or _eqstr(
                name, "Midpoint to endpoint factors"):
            continue
        _read_mid_points(wb[name], records)

    return dfutil.data_frame(records)


def _read_endpoints(file: str) -> pd.DataFrame:
    log.info(f"reading endpoint factors from file {file}")
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    endpoint = pd.DataFrame()
    endpoints = {}
    perspectives = ["I", "H", "E"]
    indicator = ""
    indicator_unit = ""
    sheet = wb['Midpoint to endpoint factors']
    start_row, data_col, with_perspectives = _find_data_start(sheet)
    # impact categories in column 1
    flow_col = 0

    endpoint_factor_count = 0
    for row in sheet.iter_rows(min_row=start_row):
        indicator = xls.cell_str(row[flow_col])
        indicator_unit = xls.cell_str(row[flow_col+1])
        for i in range(0, 3):
            val = xls.cell_f64(row[data_col + i])
            if val == 0.0:
                continue
            endpoints['Method'] = "ReCiPe 2016 - Midpoint/" + perspectives[i]
            endpoints['EndpointMethod'] = "ReCiPe 2016 - Endpoint/" + perspectives[i]
            endpoints['EndpointIndicator'] = indicator
            endpoints['EndpointUnit'] = indicator_unit
            endpoints['EndpointConversion'] = val
            endpoint = pd.concat(
                [endpoint, pd.DataFrame.from_dict([endpoints])],
                ignore_index=True)
            endpoint_factor_count += 1
    log.debug("extracted %i endpoint factors", endpoint_factor_count)

    log.info("processing endpoint factors")
    endpoint.loc[endpoint['EndpointUnit'].str.contains('daly', case=False), 'EndpointUnit'] = 'DALY'
    endpoint.loc[endpoint['EndpointUnit'].str.contains('species', case=False), 'EndpointUnit'] = 'species-year'
    endpoint.loc[endpoint['EndpointUnit'].str.contains('USD', case=False), 'EndpointUnit'] = 'USD2013'

    endpoint_map = pd.read_csv(datapath + 'ReCiPe2016_endpoint_to_midpoint.csv')
    endpoint = endpoint.merge(endpoint_map, how="left", on='EndpointIndicator')

    # split into two dataframes
    endpoint_by_flow = endpoint[endpoint['FlowFlag'] == 1]
    endpoint_by_flow = endpoint_by_flow.drop(columns='FlowFlag')
    endpoint_by_flow.rename(columns={'EndpointIndicator': 'Flowable'},
                            inplace=True)
    endpoint = endpoint[endpoint['FlowFlag'].isna()]
    endpoint = endpoint.drop(columns='FlowFlag')
    # return endpoint and endpoint by flow
    return endpoint, endpoint_by_flow


def _read_mid_points(sheet: openpyxl.worksheet.worksheet.Worksheet,
                     records: list):
    log.debug("try to read midpoint factors from sheet %s", sheet.title)

    start_row, data_col, with_perspectives = _find_data_start(sheet)
    if start_row < 0:
        log.debug("could not find a value column in sheet %s", sheet.title)
        return

    flow_col = _find_flow_column(sheet)
    if flow_col < 0:
        return

    cas_col = _find_cas_column(sheet)
    indicator_unit, flow_unit, unit_col = _determine_units(sheet)
    compartment, compartment_col = _determine_compartments(sheet)

    perspectives = ["I", "H", "E"]
    factor_count = 0
    for row in sheet.iter_rows(min_row=start_row):
        if compartment_col > -1:
            compartment = xls.cell_str(row[compartment_col])
        if compartment in contexts:
            compartment = contexts[compartment]
        if unit_col > -1:
            flow_unit = xls.cell_str(row[unit_col])
            if "/" in flow_unit:
                flow_unit = flow_unit.split("/")[1].strip()
        cas = ""
        if cas_col > -1:
            cas = format_cas(xls.cell_f64(row[cas_col]))

        if with_perspectives:
            for i in range(0, 3):
                val = xls.cell_f64(row[data_col + i])
                if val == 0.0:
                    continue
                dfutil.record(records,
                              method="ReCiPe 2016 - Midpoint/" + perspectives[i],
                              indicator=sheet.title,
                              indicator_unit=indicator_unit,
                              flow=xls.cell_str(row[flow_col]),
                              flow_category=compartment,
                              flow_unit=flow_unit,
                              cas_number=cas,
                              factor=val)
                factor_count += 1
        else:
            val = xls.cell_f64(row[data_col])
            if val == 0.0:
                continue
            for p in perspectives:
                dfutil.record(records,
                              method="ReCiPe 2016 - Midpoint/" + p,
                              indicator=sheet.title,
                              indicator_unit=indicator_unit,
                              flow=xls.cell_str(row[flow_col]),
                              flow_category=compartment,
                              flow_unit=flow_unit,
                              cas_number=cas,
                              factor=val)
                factor_count += 1
    log.debug("extracted %i factors", factor_count)


def _find_data_start(sheet: openpyxl.worksheet.worksheet.Worksheet) -> (int, int, bool):
    for row in sheet.iter_rows():
        for cell in row:
            s = xls.cell_str(cell)
            if s is None or s == "":
                continue
            if _eqstr(s, "I") or _containstr(s, "Individualist") or _containstr(s, "Individualistic"):
                return cell.row + 1, cell.column - 1, True
            if _eqstr(s, "all perspectives"):
                return cell.row + 1, cell.column - 1, False
    return -1, -1


def _find_flow_column(sheet: openpyxl.worksheet.worksheet.Worksheet) -> int:
    if _containstr(sheet.title, "land", "occupation"):
        ncol = 1
        return ncol
    ncol = -1
    for row in sheet.iter_rows():
        for cell in row:
            s = xls.cell_str(cell)
            if _containstr(s, "name") or _containstr(s, "substance"):
                ncol = cell.column - 1
                log.debug("identified column %i %s for flow names", ncol, s)
                break
    if ncol < 0:
        log.debug("no 'name' column in %s, take col=0 for that", sheet.title)
        ncol = 0
    return ncol


def _find_cas_column(sheet: openpyxl.worksheet.worksheet.Worksheet) -> int:
    ccol = -1
    for row in sheet.iter_rows():
        for cell in row:
            s = xls.cell_str(cell)
            if _eqstr(s, "cas"):
                ccol = cell.column - 1
                log.debug("identified column %i %s for CAS numbers", ccol, s)
                break
    return ccol


def _determine_units(sheet: openpyxl.worksheet.worksheet.Worksheet) -> (str, str, int):
    indicator_unit = "?"
    flow_unit = "?"
    unit_col = -1
    row, col, _ = _find_data_start(sheet)
    row -= 2

    if row > 0:
        s = xls.cell_str(sheet.cell(row=row, column=col + 1))
        if s is not None and s != "":
            if "/" in s:
                parts = s.strip(" ()").split("/")
                indicator_unit = parts[0].strip()
                flow_unit = parts[1].strip()
            else:
                indicator_unit = s.strip()

    for count, row in enumerate(sheet.iter_rows()):
        if count > 5:
            break
        for cell in row:
            s = xls.cell_str(cell)
            if _eqstr(s, "Unit"):
                unit_col = cell.column - 1
                break

    if indicator_unit != "?":
        log.debug("determined indicator unit: %s", indicator_unit)
    elif _containstr(sheet.title, "land", "transformation"):
        log.debug("unknown indicator unit; assuming it is m2")
        indicator_unit = "m2"
    elif _containstr(sheet.title, "land", "occupation"):
        log.debug("unknown indicator unit; assuming it is m2*a")
        indicator_unit = "m2*a"
    elif _containstr(sheet.title, "water", "consumption"):
        log.debug("unknown indicator unit; assuming it is m3")
        indicator_unit = "m3"
    else:
        log.debug("unknown indicator unit")

    if _containstr(flow_unit, "kg"):
        flow_unit = "kg"

    if unit_col > -1:
        log.debug("take units from column %i", unit_col)
    elif flow_unit != "?":
        log.debug("determined flow unit: %s", flow_unit)
    elif _containstr(sheet.title, "land", "transformation"):
        log.debug("unknown flow unit; assume it is m2")
        flow_unit = "m2"
    elif _containstr(sheet.title, "land", "occupation"):
        log.debug("unknown flow unit; assuming it is m2*a")
        flow_unit = "m2*a"
    elif _containstr(sheet.title, "water", "consumption"):
        log.debug("unknown flow unit; assuming it is m3")
        flow_unit = "m3"
    else:
        log.debug("unknown flow unit; assuming it is 'kg'")
        flow_unit = "kg"

    return indicator_unit, flow_unit, unit_col


def _determine_compartments(sheet: openpyxl.worksheet.worksheet.Worksheet) -> (str, int):
    compartment_col = -1
    for count, row in enumerate(sheet.iter_rows()):
        if count > 5:
            break
        for cell in row:
            s = xls.cell_str(cell)
            if _containstr(s, "compartment") or _containstr(
                    s, "name", "in", "ReCiPe"):
                compartment_col = cell.column - 1
                break

    if compartment_col > -1:
        log.debug("found compartment column %i", compartment_col)
        return "", compartment_col

    elif _containstr(sheet.title, "global", "warming") \
            or _containstr(sheet.title, "ozone") \
            or _containstr(sheet.title, "particulate") \
            or _containstr(sheet.title, "acidification"):
        log.debug("no compartment column; assuming 'air'")
        return "air", -1

    elif _containstr(sheet.title, "mineral", "resource", "scarcity"):
        log.debug("no compartment column; assuming 'resource/ground'")
        return "resource/ground", -1

    elif _containstr(sheet.title, "fossil", "resource", "scarcity"):
        log.debug("no compartment column; assuming 'resource'")
        return "resource", -1

    if _containstr(sheet.title, "water", "consumption"):
        log.debug("no compartment column; assuming 'resource/fresh water'")
        return "resource/fresh water", -1

    log.debug("no compartment column")
    return "", -1


def _eqstr(s1: str, s2: str) -> bool:
    if s1 is None or s2 is None:
        return False
    return s1.strip().lower() == s2.strip().lower()


def _containstr(s: str, *words) -> bool:
    if s is None:
        return False
    base = s.lower()
    for w in words:
        if not isinstance(w, str):
            return False
        if w.lower().strip() not in base:
            return False
    return True
