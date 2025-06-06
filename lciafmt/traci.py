# traci.py (lciafmt)
# !/usr/bin/env python3
# coding=utf-8
"""
This module contains functions needed to compile LCIA methods from EPA's
Tool for Reduction and Assessment of Chemicals and Other Environmental
Impacts (TRACI)
"""

import pandas as pd
import openpyxl

import lciafmt
import lciafmt.cache as cache
import lciafmt.df as dfutil
import lciafmt.xls as xls

from lciafmt.util import log, aggregate_factors_for_primary_contexts, format_cas,\
    datapath


flowables_replace = pd.read_csv(datapath / 'TRACI_2.1_replacement.csv')
flowables_split = pd.read_csv(datapath / 'TRACI_2.1_split.csv')


def get(method, add_factors_for_missing_contexts=True, file=None,
        url=None) -> pd.DataFrame:
    """Generate a method for TRACI in standard format.

    :param add_factors_for_missing_contexts: bool, if True generates average
        factors for unspecified contexts
    :param file: str, alternate filepath for method, defaults to file stored
        in cache
    :param url: str, alternate url for method, defaults to url in method config
    :return: DataFrame of method in standard format
    """
    log.info("getting method TRACI")
    method_meta = method.get_metadata()
    f = file
    if f is None:
        f = _get_file(method_meta, url)
    df = _read(f)
    if add_factors_for_missing_contexts:
        log.info("adding average factors for primary contexts")
        df = aggregate_factors_for_primary_contexts(df)

    log.info("handling manual replacements")
    """ due to substances listed more than once with different names
    this replaces all instances of the Original Flowable with a New Flowable
    based on a csv input file, otherwise zero values for CFs will override
    when there are duplicate names"""
    for index, row in flowables_replace.iterrows():
        orig = row['Original Flowable']
        new = row['New Flowable']
        df['Flowable'] = df['Flowable'].replace(orig, new)

    """due to substances listed more than once with the same name but different
    CAS this replaces all instances of the Original Flowable with a New Flowable
    based on a csv input file according to the CAS"""
    for index, row in flowables_split.iterrows():
        cas = row['CAS']
        new = row['New Flowable']
        df.loc[df['CAS No'] == cas, 'Flowable'] = new

    length = len(df)
    df.drop_duplicates(keep='first', inplace=True)
    length = length - len(df)
    log.info(f"{length} duplicate entries removed")
    
    """add eutrophication updates 
    the function _read_eutro is a function to read the raw data from the new
    eutrophication updates
    """
    if 'eutro_url' in method_meta:
        log.info("getting Eutrophication updates")
        f = cache.get_or_download(file=method_meta['eutro_file'],
                                  url=method_meta['eutro_url'])
        df_eutro = _read_eutro(f)
        frames = [df.query('Indicator != "Eutrophication"'), df_eutro]
        df = pd.concat(frames)
    df['Method'] = method_meta.get('name')

    return df

def _get_file(method_meta, url=None):
    fname = "traci_2.1.xlsx"
    if url is None:
        url = method_meta['url']
    f = cache.get_or_download(fname, url)
    return f

def _read(xls_file: str) -> pd.DataFrame:
    """Read the data from Excel with given path into a DataFrame."""
    log.info(f"read TRACI from file {xls_file}")
    wb = openpyxl.load_workbook(xls_file, read_only=True, data_only=True)
    sheet = wb["Substances"]
    categories = {}
    max_col = sheet.max_column
    for count, cell in enumerate(list(sheet.rows)[0]):
        name = xls.cell_str(cell)
        if name == "":
            break
        cat_info = _category_info(name)
        if cat_info is not None:
            categories[count+1] = cat_info

    records = []
    for row in sheet.iter_rows(min_row=2):
        flow = xls.cell_str(row[2])
        if flow == "":
            break
        cas = format_cas((row[1]).value)
        for col in range(4, max_col):
            cat_info = categories.get(col)
            if cat_info is None:
                continue
            factor = xls.cell_f64(row[col-1])
            if factor == 0.0:
                continue
            dfutil.record(records,
                          indicator=cat_info[0],
                          indicator_unit=cat_info[1],
                          flow=flow,
                          flow_category=cat_info[2],
                          flow_unit=cat_info[3],
                          cas_number=cas,
                          factor=factor)
    wb.close()
    return dfutil.data_frame(records)


def _category_info(c: str):
    """Parse category field into metdata.

    Get the meta data which are encoded in the category name. It returns
    a tuple (indicator, indicator unit, flow category, flow unit) for the
    given category name. If it is an unknown category, `None` is returned.
    """
    if c == "Global Warming Air (kg CO2 eq / kg substance)":
        return "Global warming", "kg CO2 eq", "air", "kg"

    if c == "Global Climate Air (kg CO2 eq / kg substance)":
        return "Global warming", "kg CO2 eq", "air", "kg"

    if c == "Acidification Air (kg SO2 eq / kg substance)":
        return "Acidification", "kg SO2 eq", "air", "kg"

    if c == "HH Particulate Air (PM2.5 eq / kg substance)":
        return ("Human health - particulate matter", "PM 2.5 eq",
                "air", "kg")

    if c == "Eutrophication Air (kg N eq / kg substance)":
        return "Eutrophication", "kg N eq", "air", "kg"

    if c == "Eutrophication Water (kg N eq / kg substance)":
        return "Eutrophication", "kg N eq", "water", "kg"

    if c == "Ozone Depletion Air (kg CFC-11 eq / kg substance)":
        return "Ozone depletion", "kg CFC-11 eq", "air", "kg"

    if c == "Smog Air (kg O3 eq / kg substance)":
        return "Smog formation", "kg O3 eq", "air", "kg"

    if c == "Ecotox. CF [CTUeco/kg], Em.airU, freshwater":
        return "Freshwater ecotoxicity", "CTUeco", "air/urban", "kg"

    if c == "Ecotox. CF [CTUeco/kg], Em.airC, freshwater":
        return "Freshwater ecotoxicity", "CTUeco", "air/rural", "kg"

    if c == "Ecotox. CF [CTUeco/kg], Em.fr.waterC, freshwater":
        return "Freshwater ecotoxicity", "CTUeco", "water/freshwater", "kg"

    if c == "Ecotox. CF [CTUeco/kg], Em.sea waterC, freshwater":
        return "Freshwater ecotoxicity", "CTUeco", "water/sea water", "kg"

    if c == "Ecotox. CF [CTUeco/kg], Em.nat.soilC, freshwater":
        return "Freshwater ecotoxicity", "CTUeco", "soil/natural", "kg"

    if c == "Ecotox. CF [CTUeco/kg], Em.agr.soilC, freshwater":
        return "Freshwater ecotoxicity", "CTUeco", "soil/agricultural", "kg"

    if c == "Human health CF  [CTUcancer/kg], Emission to urban air, cancer":
        return "Human health - cancer", "CTUcancer", "air/urban", "kg"

    if c == "Human health CF  [CTUnoncancer/kg], Emission to urban air, non-canc.":
        return "Human health - non-cancer", "CTUnoncancer", "air/urban", "kg"

    if c == "Human health CF  [CTUcancer/kg], Emission to cont. rural air, cancer":
        return "Human health - cancer", "CTUcancer", "air/rural", "kg"

    if c == "Human health CF  [CTUnoncancer/kg], Emission to cont. rural air, non-canc.":
        return "Human health - non-cancer", "CTUnoncancer", "air/rural", "kg"

    if c == "Human health CF  [CTUcancer/kg], Emission to cont. freshwater, cancer":
        return "Human health - cancer", "CTUcancer", "water/freshwater", "kg"

    if c == "Human health CF  [CTUnoncancer/kg], Emission to cont. freshwater, non-canc.":
        return "Human health - non-cancer", "CTUnoncancer", "water/freshwater", "kg"

    if c == "Human health CF  [CTUcancer/kg], Emission to cont. sea water, cancer":
        return "Human health - cancer", "CTUcancer", "water/sea water", "kg"

    if c == "Human health CF  [CTUnoncancer/kg], Emission to cont. sea water, non-canc.":
        return "Human health - non-cancer", "CTUnoncancer", "water/sea water", "kg"

    if c == "Human health CF  [CTUcancer/kg], Emission to cont. natural soil, cancer":
        return "Human health - cancer", "CTUcancer", "soil/natural", "kg"

    if c == "Human health CF  [CTUnoncancer/kg], Emission to cont. natural soil, non-canc.":
        return "Human health - non-cancer", "CTUnoncancer", "soil/natural", "kg"

    if c == "Human health CF  [CTUcancer/kg], Emission to cont. agric. Soil, cancer":
        return "Human health - cancer", "CTUcancer", "soil/agricultural", "kg"

    if c == "Human health CF  [CTUnoncancer/kg], Emission to cont. agric. Soil, non-canc.":
        return "Human health - non-cancer", "CTUnoncancer", "soil/agricultural", "kg"

def _read_eutro(xls_file: str) -> pd.DataFrame:
    """
    Logic used for selecting US data (max 15 per region):
    |               | Comp_Air | Comp_Fw | Comp_Soil | Comp_LME |
    |---------------|----------|---------|-----------|----------|
    | Flow_N        | n/a      | NonAg   | Agric     | NonAg    |
    | Flow_NH3 as N | All      | n/a     | n/a       | n/a      |
    | Flow_NOx as N | All      | n/a     | n/a       | n/a      |
    | Flow_P        | n/a      | All     | All       | n/a      |
    * skip Agric & NonAg

    Logic used for selecting global data (max 15 per region):
    |               | Comp_Air | Comp_Fw        | Comp_Soil | Comp_LME |
    |---------------|----------|----------------|-----------|----------|
    | Flow_N        | n/a      | Genrl == NonAg | Agric     | NonAg    |
    | Flow_NH3 as N | All      | n/a            | n/a       | n/a      |
    | Flow_NOx as N | All      | n/a            | n/a       | n/a      |
    | Flow_P        | n/a      | All            | All       | n/a      |

    """
    context_dict = {'Comp_Fw': 'freshwater',
                    'Comp_Air': 'air',
                    'Comp_Soil': 'soil',
                    'Comp_LME': 'marine'}
    compartment_dict = {'Genrl': 'unspecified',
                        'Agric': 'rural',
                        'NonAg': 'urban'}
    log.info(f"read Eutrophication category from file {xls_file}")
    source_df = pd.read_excel(xls_file, sheet_name="S5. Raw Data")
    records = []
    flow_category=[]
    for i, row in source_df.iterrows():
        sector = row['Sector']
        flow = row['Flowable']
        compartment = row['Emit Compartment']
        flow_category = context_dict.get(compartment, "n/a")
        if flow_category == "soil" and flow == "Flow_P":
            flow_category = f'{flow_category} (P)'
            # required to enable distinct mappings to ground for these flows
        aggregation = row['Aggregation Target']
        region_id = str(row['Target ID'])
        if aggregation in ("US_Nation", "US_States", "US_Counties"):
            if aggregation == "US_Nation":
                region = "00000"
            elif len(region_id) < 3:
                region = region_id.ljust(5, '0')
            else:
                region = region_id.rjust(5, '0')

        elif (aggregation in ("World", "Countries")):
            region = row['Name']
            if region == "United States":
                ## ^^ Skip US as country in favor of aggregation == "US_Nation"
                continue
            elif region == "Russian Federation" and region_id == "254":
                ## Two entries for Russian Federation, 254 is a very small island
                continue
            if sector == "Genrl" and flow == "Flow_N":
                ## Drops duplicate factors for Flow_N Comp_Fw
                continue
        else:
            # Ignore aggregation == "Continents"
            continue
        if flow != "Flow_N":
            flow_category = f'{flow_category}/{compartment_dict.get(sector)}'

        factor = row['Average Target Value']
        indicator = ("Eutrophication (Freshwater)" if flow == "Flow_P"
                     else "Eutrophication (Marine)")
        unit = ("kg P eq" if indicator == "Eutrophication (Freshwater)"
                else "kg N eq")

        dfutil.record(records,
                      indicator=indicator,
                      indicator_unit=unit,
                      flow=flow,
                      flow_category=flow_category,
                      flow_unit="kg",
                      factor=factor,
                      location=region)

        if aggregation == "World":
        # openLCA requires a factor without location for use by default
            dfutil.record(records,
                          indicator=indicator,
                          indicator_unit=unit,
                          flow=flow,
                          flow_category=flow_category,
                          flow_unit="kg",
                          factor=factor,
                          location="")

    df = dfutil.data_frame(records)

    # Resolve duplicate factors for a single location
    cols_to_keep = [c for c in df.columns if
                    c not in ('Characterization Factor')]
    duplicates = df[df.duplicated(subset=cols_to_keep, keep=False)]
    ## United States Minor Outlying Islands and Jan Mayen
    # are unexplicably shown multiple times with different location IDs.
    # Average those factors together.
    df2 = (df.groupby(cols_to_keep, as_index=False)
             .agg({'Characterization Factor': 'mean'}))
    log.debug(f'{len(duplicates)} duplicate locations consolidated to '
             f'{(len(duplicates)-(len(df)-len(df2)))}')

    return df2


#%%
if __name__ == "__main__":
    from lciafmt.util import store_method
    method = lciafmt.Method.TRACI2_2
    df = get(method)
    mapping = method.get_metadata()['mapping']
    mapped_df = lciafmt.map_flows(df, system=mapping)
    store_method(mapped_df, method)
